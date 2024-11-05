import tkinter as tk
from tkinter import messagebox, font, ttk
import sqlite3
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# Database Setup
def connect_db():
    conn = sqlite3.connect("inventory.db")
    cursor = conn.cursor()
    cursor.execute("""CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT NOT NULL CHECK(role IN ('Admin', 'User'))
        )""")
    cursor.execute("""CREATE TABLE IF NOT EXISTS products (
            product_id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            category TEXT NOT NULL,
            price REAL NOT NULL,
            stock_quantity INTEGER NOT NULL
        )""")
    try:
        cursor.execute("INSERT INTO users (username, password, role) VALUES ('admin', 'admin123', 'Admin')")
        cursor.execute("INSERT INTO users (username, password, role) VALUES ('user', 'user123', 'User')")
    except sqlite3.IntegrityError:
        pass
    conn.commit()
    return conn

def authenticate_user(username, password):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT role FROM users WHERE username=? AND password=?", (username, password))
    user = cursor.fetchone()
    conn.close()
    return user[0] if user else None

# Admin Panel - Add Product Functionality
def add_product(name, category, price, stock, tree):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO products (name, category, price, stock_quantity) VALUES (?, ?, ?, ?)",
                   (name, category, float(price), int(stock)))
    conn.commit()
    conn.close()
    messagebox.showinfo("Success", "Product added successfully")
    log_to_excel(name, category, price, stock)
    refresh_product_table(tree)  # Refresh the product table after adding

def log_to_excel(name, category, price, stock):
    filename = "product_log.xlsx"
    file_exists = os.path.isfile(filename)
    if not file_exists:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Product Log"
        sheet.append(["Date", "Product Name", "Category", "Price", "Stock Quantity"])
    else:
        workbook = load_workbook(filename)
        sheet = workbook.active
    current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append([current_date, name, category, float(price), int(stock)])
    workbook.save(filename)

# Admin Panel - Edit Products Functionality
def get_products():
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM products")
    products = cursor.fetchall()
    conn.close()
    return products

def update_stock(product_id, new_stock, tree):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE products SET stock_quantity=? WHERE product_id=?", (new_stock, product_id))
    conn.commit()
    conn.close()
    messagebox.showinfo("Success", "Stock updated successfully")
    refresh_product_table(tree)  # Refresh the product table after updating stock

def delete_product(product_id, tree):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM products WHERE product_id=?", (product_id,))
    conn.commit()
    conn.close()
    messagebox.showinfo("Success", "Product deleted successfully")
    refresh_product_table(tree)  # Refresh the product table after deleting

# User Panel - Generate Bill Functionality
def generate_bill(items):
    total = sum(item['price'] * item['quantity'] for item in items)
    bill_window = tk.Toplevel()
    bill_window.title("Generated Bill")

    columns = ("Item", "Quantity", "Price", "Total")
    tree = ttk.Treeview(bill_window, columns=columns, show='headings')
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, anchor="center")
    tree.pack(expand=True, fill="both", padx=10, pady=10)

    for item in items:
        total_price = item['price'] * item['quantity']
        tree.insert("", "end", values=(item['name'], item['quantity'], f"${item['price']:.2f}", f"${total_price:.2f}"))

    total_label = tk.Label(bill_window, text=f"Total Amount: ${total:.2f}", font=("Arial", 14), fg="blue")
    total_label.pack(pady=10)

# Refresh Product Table
def refresh_product_table(tree):
    for item in tree.get_children():
        tree.delete(item)
    products = get_products()
    for product in products:
        tree.insert("", "end", values=product)

# Search Product
def search_product(product_id, tree):
    for item in tree.get_children():
        tree.delete(item)
    products = get_products()
    found = False
    for product in products:
        if product[0] == product_id:
            tree.insert("", "end", values=product, tags=('highlight',))
            found = True
        else:
            tree.insert("", "end", values=product)
    if not found:
        messagebox.showinfo("Search Result", "Product ID not found.")

# Login Page
def create_login_page(root):
    root.title("Login")
    root.geometry("400x300")
    
    heading_font = font.Font(family="Bebas Neue", size=24, weight="bold")
    tk.Label(root, text="User Login", font=heading_font, fg="black").pack(pady=20)

    tk.Label(root, text="Username", font=("Arial", 12)).pack(pady=(10, 5))
    username_entry = ttk.Entry(root, width=30, font=("Arial", 12))
    username_entry.pack()

    tk.Label(root, text="Password", font=("Arial", 12)).pack(pady=(10, 5))
    password_entry = ttk.Entry(root, show="*", width=30, font=("Arial", 12))
    password_entry.pack()

    def login():
        username = username_entry.get()
        password = password_entry.get()
        role = authenticate_user(username, password)
        if role:
            root.destroy()
            new_root = tk.Tk()
            if role == 'Admin':
                create_admin_panel(new_root)
            else:
                create_user_panel(new_root)
            new_root.mainloop()
        else:
            messagebox.showerror("Login Failed", "Invalid username or password.")

    login_button = ttk.Button(root, text="Login", command=login, width=15)
    login_button.pack(pady=20)

# Admin Panel
def create_admin_panel(root):
    root.title("Admin Panel")
    root.geometry("800x600")

    notebook = ttk.Notebook(root)

    # Add Products Tab
    add_tab = ttk.Frame(notebook)
    notebook.add(add_tab, text="Add Products")
    tk.Label(add_tab, text="Add New Product", font=("Arial", 14)).pack(pady=10)

    tk.Label(add_tab, text="Product Name").pack(pady=5)
    name_entry = tk.Entry(add_tab, width=30)
    name_entry.pack()

    tk.Label(add_tab, text="Category").pack(pady=5)
    category_entry = tk.Entry(add_tab, width=30)
    category_entry.pack()

    tk.Label(add_tab, text="Price").pack(pady=5)
    price_entry = tk.Entry(add_tab, width=30)
    price_entry.pack()

    tk.Label(add_tab, text="Stock Quantity").pack(pady=5)
    stock_entry = tk.Entry(add_tab, width=30)
    stock_entry.pack()

    # Pass the tree view to add_product function
    add_button = ttk.Button(add_tab, text="Add Product", command=lambda: add_product(
        name_entry.get(), category_entry.get(), price_entry.get(), stock_entry.get(), tree))
    add_button.pack(pady=10)

    # Edit Products Tab
    edit_tab = ttk.Frame(notebook)
    notebook.add(edit_tab, text="Edit Products")
    tk.Label(edit_tab, text="Edit Product Stock", font=("Arial", 14)).pack(pady=10)

    tk.Label(edit_tab, text="Product ID").pack(pady=5)
    product_id_entry = tk.Entry(edit_tab, width=30)  # Entry for Product ID
    product_id_entry.pack(pady=5)

    tk.Label(edit_tab, text="New Stock Quantity").pack(pady=5)
    new_stock_entry = tk.Entry(edit_tab, width=30)
    new_stock_entry.pack()

    # Pass the tree view to update_stock function
    update_button = ttk.Button(edit_tab, text="Update Stock", command=lambda: update_stock(
        int(product_id_entry.get()), new_stock_entry.get(), tree))  # Use the Product ID from the new entry
    update_button.pack(pady=10)

    delete_button = ttk.Button(edit_tab, text="Delete Product", command=lambda: delete_product(
        int(product_id_entry.get()), tree))  # Use the Product ID from the new entry
    delete_button.pack(pady=10)

    # View Products Tab
    view_tab = ttk.Frame(notebook)
    notebook.add(view_tab, text="View Products")
    tk.Label(view_tab, text="Products List", font=("Arial", 14)).pack(pady=10)

    columns = ("Product ID", "Name", "Category", "Price", "Stock Quantity")
    tree = ttk.Treeview(view_tab, columns=columns, show='headings')
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, anchor="center")
    tree.pack(expand=True, fill="both", padx=10, pady=10)

    refresh_product_table(tree)

    notebook.pack(expand=True, fill="both", padx=10, pady=10)

    # Logout Button
    ttk.Button(root, text="Logout", command=lambda: logout(root), width=10).pack(pady=10)

# User Panel
def create_user_panel(root):
    root.title("User Panel")
    root.geometry("800x600")

    tk.Label(root, text="User Panel", font=("Arial", 24)).pack(pady=20)

    # Bill Generation Functionality
    items = []  # Store selected items for billing

    def add_item():
        product_id = int(product_id_entry.get())
        quantity = int(quantity_entry.get())
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("SELECT name, price FROM products WHERE product_id=?", (product_id,))
        product = cursor.fetchone()
        if product:
            name, price = product
            items.append({'name': name, 'price': price, 'quantity': quantity})
            refresh_billing_list()
        else:
            messagebox.showerror("Error", "Product not found.")

    def refresh_billing_list():
        for item in billing_tree.get_children():
            billing_tree.delete(item)
        for item in items:
            billing_tree.insert("", "end", values=(item['name'], item['quantity'], f"${item['price']:.2f}", f"${item['price'] * item['quantity']:.2f}"))

    def generate_bill_button():
        if items:
            generate_bill(items)
            items.clear()
            refresh_billing_list()
        else:
            messagebox.showerror("Error", "No items added to bill.")

    billing_frame = ttk.Frame(root)
    billing_frame.pack(pady=10)

    tk.Label(billing_frame, text="Product ID").pack(side=tk.LEFT)
    product_id_entry = tk.Entry(billing_frame, width=10)
    product_id_entry.pack(side=tk.LEFT)

    tk.Label(billing_frame, text="Quantity").pack(side=tk.LEFT)
    quantity_entry = tk.Entry(billing_frame, width=10)
    quantity_entry.pack(side=tk.LEFT)

    add_button = ttk.Button(billing_frame, text="Add to Bill", command=add_item)
    add_button.pack(side=tk.LEFT)

    generate_bill_btn = ttk.Button(root, text="Generate Bill", command=generate_bill_button)
    generate_bill_btn.pack(pady=10)

    billing_columns = ("Item", "Quantity", "Price", "Total")
    billing_tree = ttk.Treeview(root, columns=billing_columns, show='headings')
    for col in billing_columns:
        billing_tree.heading(col, text=col)
    billing_tree.pack(expand=True, fill="both", padx=10, pady=10)

    # Logout Button
    ttk.Button(root, text="Logout", command=lambda: logout(root), width=10).pack(pady=10)

# Logout Function
def logout(root):
    root.destroy()
    new_root = tk.Tk()
    create_login_page(new_root)
    new_root.mainloop()

# Main Function
if __name__ == "__main__":
    root = tk.Tk()
    create_login_page(root)
    root.mainloop()
