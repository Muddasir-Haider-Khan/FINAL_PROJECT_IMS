from flask import Flask, render_template, request, redirect, url_for, flash, session
import sqlite3
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

app = Flask(__name__)
app.secret_key = 'your_secret_key'

# Database Setup
def connect_db():
    conn = sqlite3.connect("inventory.db")
    return conn

def init_db():
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("""CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT NOT NULL CHECK(role IN ('Admin', 'User'))
        )""")
    cursor.execute("""CREATE TABLE IF NOT EXISTS products (
            product_id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            category TEXT NOT NULL,
            price REAL NOT NULL,
            stock_quantity INTEGER NOT NULL
        )""")
    # Default admin and user
    try:
        cursor.execute("INSERT INTO users (username, password, role) VALUES ('admin', 'admin123', 'Admin')")
        cursor.execute("INSERT INTO users (username, password, role) VALUES ('user', 'user123', 'User')")
    except sqlite3.IntegrityError:
        pass
    conn.commit()
    conn.close()

# User authentication
def authenticate_user(username, password):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT role FROM users WHERE username=? AND password=?", (username, password))
    user = cursor.fetchone()
    conn.close()
    return user[0] if user else None

# Add product to the database
def add_product(name, category, price, stock):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO products (name, category, price, stock_quantity) VALUES (?, ?, ?, ?)",
                   (name, category, float(price), int(stock)))
    conn.commit()
    conn.close()
    log_to_excel(name, category, price, stock)

# Log product details to an Excel file
def log_to_excel(name, category, price, stock):
    filename = "product_log.xlsx"
    file_exists = os.path.isfile(filename)

    if not file_exists:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Product Log"
        sheet.append(["Date", "Product Name", "Category", "Price", "Stock Quantity"])
    else:
        workbook = load_workbook(filename)
        sheet = workbook.active

    current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append([current_date, name, category, float(price), int(stock)])
    workbook.save(filename)

# Get all products from the database
def get_products():
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM products")
    products = cursor.fetchall()
    conn.close()
    return products

# Delete a product from the database
def delete_product(product_id):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM products WHERE product_id=?", (product_id,))
    conn.commit()
    conn.close()

# Update stock quantity for a product
def update_stock(product_id, new_quantity):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE products SET stock_quantity=? WHERE product_id=?", (new_quantity, product_id))
    conn.commit()
    conn.close()

@app.route('/')
def login():
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def do_login():
    username = request.form['username']
    password = request.form['password']
    role = authenticate_user(username, password)
    if role:
        session['username'] = username
        session['role'] = role
        return redirect(url_for('admin_panel' if role == 'Admin' else 'user_panel'))
    else:
        flash('Invalid username or password.')
        return redirect(url_for('login'))

@app.route('/admin_panel')
def admin_panel():
    if 'role' not in session or session['role'] != 'Admin':
        return redirect(url_for('login'))
    products = get_products()
    return render_template('admin_panel.html', products=products)

@app.route('/add_product', methods=['POST'])
def add_new_product():
    name = request.form['name']
    category = request.form['category']
    price = request.form['price']
    stock = request.form['stock']
    add_product(name, category, price, stock)
    flash('Product added successfully!')
    return redirect(url_for('admin_panel'))

@app.route('/delete_product/<int:product_id>', methods=['POST'])
def delete_product_route(product_id):
    delete_product(product_id)
    flash('Product deleted successfully!')
    return redirect(url_for('admin_panel'))

@app.route('/user_panel')
def user_panel():
    if 'role' not in session or session['role'] != 'User':
        return redirect(url_for('login'))
    products = get_products()
    return render_template('user_panel.html', products=products)

@app.route('/generate_bill', methods=['POST'])
def generate_bill():
    items = request.form.getlist('items')
    quantities = request.form.getlist('quantities')

    bill = []
    total_amount = 0
    for item, qty in zip(items, quantities):
        if qty and item:
            product_id = int(item)
            quantity = int(qty)
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("SELECT name, price FROM products WHERE product_id=?", (product_id,))
            product = cursor.fetchone()
            conn.close()
            if product:
                name, price = product
                total_price = price * quantity
                total_amount += total_price
                bill.append({"name": name, "quantity": quantity, "price": price, "total": total_price})

    return render_template('bill.html', bill=bill, total_amount=total_amount)

@app.route('/logout')
def logout():
    session.pop('username', None)
    session.pop('role', None)
    return redirect(url_for('login'))

if __name__ == '__main__':
    try:
        init_db()  # Initialize the database
        app.run(debug=True, port=5000, host='0.0.0.0')
    except Exception as e:
        print(f"An error occurred: {e}")